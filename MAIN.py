import imaplib
import email
import os
import re
import sqlite3
import threading
from email.header import decode_header
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from tkinter import Tk, Label, Entry, Button, Listbox, Scrollbar, END, messagebox, Toplevel, StringVar, OptionMenu
from tkinter import ttk

# === CONFIG ===
IMAP_HOST = "imap.gmail.com"
EMAIL_USER = "akshitamoda2006@gmail.com"
EMAIL_PASS = "ybmuduizvfeacoes"
ATTACH_DIR = "xlsx_attachments_only"
DB_PATH = "matches.db"
os.makedirs(ATTACH_DIR, exist_ok=True)

# === CLEANERS ===
def clean_subject(subject):
    decoded_parts = decode_header(subject or "")
    final_subject = ""
    for part, encoding in decoded_parts:
        if isinstance(part, bytes):
            part = part.decode(encoding or 'utf-8', errors='ignore')
        final_subject += part
    final_subject = re.sub(r'[\x00-\x1F\x7F]', '', final_subject)
    return final_subject.replace('\n', ' ').replace('\r', ' ').strip() or "NO SUBJECT"

def clean_cell_value(val):
    if val is None:
        return ''
    text = str(val)
    text = re.sub(r'[\x00-\x1F\x7F]', '', text)
    return text.replace('\n', ' ').replace('\r', ' ').strip().lower()

# === DB SETUP ===
def init_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('DROP TABLE IF EXISTS matches')
    c.execute('''CREATE TABLE matches (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        category TEXT,
        subject TEXT,
        sender TEXT,
        date TEXT,
        file_path TEXT,
        sheet TEXT,
        cell TEXT,
        cell_value TEXT
    )''')
    conn.commit()
    conn.close()

# === EMAIL PROCESSING ===
def process_emails(category_keywords):
    update_status("Connecting to mail...")
    mail = imaplib.IMAP4_SSL(IMAP_HOST)
    mail.login(EMAIL_USER, EMAIL_PASS)
    mail.select("inbox")

    result, data = mail.search(None, "ALL")
    email_ids = data[0].split()

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    total = len(email_ids)
    for idx, email_id in enumerate(email_ids, 1):
        update_status(f"Reading email {idx}/{total}...")
        progress_bar['value'] = (idx / total) * 100
        root.update_idletasks()

        result, message_data = mail.fetch(email_id, "(RFC822)")
        raw_email = message_data[0][1]
        msg = email.message_from_bytes(raw_email)

        subject = clean_subject(msg["Subject"])
        sender = msg.get("From", "Unknown")
        date = msg.get("Date", "Unknown")
        attachments = []

        if msg.is_multipart():
            for part in msg.walk():
                if "attachment" in str(part.get("Content-Disposition", "")).lower():
                    filename = part.get_filename()
                    if filename and filename.lower().endswith(".xlsx"):
                        filename = email.utils.collapse_rfc2231_value(filename)
                        safe_filename = filename.replace('/', '').replace('\\', '')
                        filepath = os.path.abspath(os.path.join(ATTACH_DIR, safe_filename))
                        with open(filepath, "wb") as f:
                            f.write(part.get_payload(decode=True))
                        attachments.append(filepath)

        for filepath in attachments:
            try:
                wb = load_workbook(filepath)
                for sheet in wb.worksheets:
                    for row in sheet.iter_rows(min_row=2, values_only=False):
                        for cell in row:
                            cell_value = clean_cell_value(cell.value)
                            for category, keywords in category_keywords.items():
                                if any(kw in cell_value for kw in keywords):
                                    c.execute("INSERT INTO matches (category, subject, sender, date, file_path, sheet, cell, cell_value) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                                              (category, subject, sender, date, filepath, sheet.title, cell.coordinate, str(cell.value)))
                                    conn.commit()
                                    break
            except Exception as e:
                print(f"Error reading {filepath}: {e}")

    conn.close()
    mail.logout()
    update_status("\u2705 Done! All data saved to database.")
    messagebox.showinfo("Done", "Email processing complete.")
    run_button.config(state="normal")
    run_button.config(text="\u25B6 Run Email Processor")
    progress_bar['value'] = 100

# === EXPORT TO EXCEL ===
def export_results_to_excel():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT DISTINCT category FROM matches")
    categories = [row[0] for row in c.fetchall()]

    for category in categories:
        c.execute("SELECT subject, sender, date, file_path, sheet, cell, cell_value FROM matches WHERE category = ?", (category,))
        rows = c.fetchall()
        if not rows:
            continue
        wb = Workbook()
        ws = wb.active
        ws.title = "Matches"
        ws.append(["Subject", "Sender", "Date", "Excel File Path", "Sheet", "Cell", "Cell Value"])
        for row in rows:
            ws.append(row)
        for col in ws.columns:
            max_length = max((len(str(cell.value)) for cell in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 5
        wb.save(f"{category}.xlsx")
    conn.close()
    messagebox.showinfo("Exported", "\u2705 All categories exported to Excel")

# === VIEWER ===
def open_viewer_window():
    viewer = Toplevel(root)
    viewer.title("Browse Matches by Category")
    viewer.geometry("900x500")

    Label(viewer, text="Select Category:").pack()
    selected_category = StringVar(viewer)
    selected_category.set("Choose...")

    dropdown = OptionMenu(viewer, selected_category, *get_categories_from_db(), command=lambda cat: show_category_data(cat, viewer))
    dropdown.pack(pady=5)

def get_categories_from_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT DISTINCT category FROM matches")
    categories = [row[0] for row in c.fetchall()]
    conn.close()
    return categories

def show_category_data(category, parent):
    for widget in parent.winfo_children():
        if isinstance(widget, ttk.Treeview):
            widget.destroy()

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT subject, sender, date, file_path, sheet, cell, cell_value FROM matches WHERE category = ?", (category,))
    rows = c.fetchall()
    conn.close()

    tree = ttk.Treeview(parent, columns=("Subject", "Sender", "Date", "File", "Sheet", "Cell", "Value"), show='headings', height=20)
    for col in tree["columns"]:
        tree.heading(col, text=col)
        tree.column(col, width=120, anchor="w")
    for row in rows:
        tree.insert("", END, values=row)
    tree.pack(pady=10)

    Button(parent, text=f"\U0001F4E4 Export '{category}' to Excel", bg="skyblue", command=lambda: export_single_category(category)).pack(pady=5)

def export_single_category(category):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT subject, sender, date, file_path, sheet, cell, cell_value FROM matches WHERE category = ?", (category,))
    rows = c.fetchall()
    conn.close()
    if not rows:
        messagebox.showinfo("No Data", f"No matches found for category: {category}")
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Matches"
    ws.append(["Subject", "Sender", "Date", "Excel File Path", "Sheet", "Cell", "Cell Value"])
    for row in rows:
        ws.append(row)
    for col in ws.columns:
        max_length = max((len(str(cell.value)) for cell in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 5
    wb.save(f"{category}.xlsx")
    messagebox.showinfo("Exported", f"\u2705 Exported to {category}.xlsx")

# === GUI LOGIC ===
category_keywords = {}

def add_category():
    cat = category_entry.get().strip()
    keys = keywords_entry.get().strip().lower().split(",")
    keys = [k.strip() for k in keys if k.strip()]
    if not cat or not keys:
        messagebox.showwarning("Invalid Input", "Please enter both category and keywords.")
        return
    category_keywords[cat] = keys
    listbox.insert(END, f"{cat}: {', '.join(keys)}")
    category_entry.delete(0, END)
    keywords_entry.delete(0, END)

def update_status(text):
    status_label.config(text=text)

def run_processor():
    if not category_keywords:
        messagebox.showerror("Error", "Add at least one category and keywords first.")
        return
    run_button.config(state="disabled")
    run_button.config(text="Processing...")
    progress_bar['value'] = 0
    update_status("Starting email scan...")
    threading.Thread(target=lambda: process_emails(category_keywords)).start()

# === GUI SETUP ===
init_db()
root = Tk()
root.title("Email Excel Parser with SQL")
root.geometry("540x520")

Label(root, text="Category Name").pack()
category_entry = Entry(root, width=50)
category_entry.pack()

Label(root, text="Keywords (comma-separated)").pack()
keywords_entry = Entry(root, width=50)
keywords_entry.pack()

Button(root, text="\u2795 Add Category", command=add_category).pack(pady=5)

scrollbar = Scrollbar(root)
scrollbar.pack(side="right", fill="y")
listbox = Listbox(root, width=60, yscrollcommand=scrollbar.set)
listbox.pack(pady=10)
scrollbar.config(command=listbox.yview)

run_button = Button(root, text="\u25B6 Run Email Processor", bg="#4CAF50", fg="white", command=run_processor)
run_button.pack(pady=5)

Button(root, text="\U0001F441 View Matches", bg="gray", fg="white", command=open_viewer_window).pack(pady=5)
Button(root, text="\U0001F4C4 Export All to Excel", bg="teal", fg="white", command=export_results_to_excel).pack(pady=5)

progress_bar = ttk.Progressbar(root, orient="horizontal", length=400, mode="determinate")
progress_bar.pack(pady=5)

status_label = Label(root, text="", fg="blue")
status_label.pack()

root.mainloop()