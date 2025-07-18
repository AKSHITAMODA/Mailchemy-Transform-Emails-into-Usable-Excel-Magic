import imaplib
import email
import os
import re
import sqlite3
from email.header import decode_header
from openpyxl import load_workbook

IMAP_HOST = "imap.gmail.com"
EMAIL_USER = "your_email@gmail.com"
EMAIL_PASS = "your_app_password"
ATTACH_DIR = "xlsx_attachments_only"
DB_PATH = "matches.db"
os.makedirs(ATTACH_DIR, exist_ok=True)

def clean_subject(subject):
    decoded_parts = decode_header(subject or "")
    final_subject = ""
    for part, encoding in decoded_parts:
        if isinstance(part, bytes):
            part = part.decode(encoding or 'utf-8', errors='ignore')
        final_subject += part
    final_subject = re.sub(r'[\x00-\x1F\x7F]', '', final_subject)
    return final_subject.replace('\n', ' ').replace('\r', ' ').strip() or "NO SUBJECT"

def clean_cell_value(val):
    if val is None:
        return ''
    text = str(val)
    text = re.sub(r'[\x00-\x1F\x7F]', '', text)
    return text.replace('\n', ' ').replace('\r', ' ').strip().lower()

def process_emails(category_keywords, update_status, progress_bar, root, run_button):
    update_status("Connecting to mail...")
    mail = imaplib.IMAP4_SSL(IMAP_HOST)
    mail.login(EMAIL_USER, EMAIL_PASS)
    mail.select("inbox")

    result, data = mail.search(None, "ALL")
    email_ids = data[0].split()

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    total = len(email_ids)
    for idx, email_id in enumerate(email_ids, 1):
        update_status(f"Reading email {idx}/{total}...")
        progress_bar['value'] = (idx / total) * 100
        root.update_idletasks()

        result, message_data = mail.fetch(email_id, "(RFC822)")
        raw_email = message_data[0][1]
        msg = email.message_from_bytes(raw_email)

        subject = clean_subject(msg["Subject"])
        sender = msg.get("From", "Unknown")
        date = msg.get("Date", "Unknown")
        attachments = []

        if msg.is_multipart():
            for part in msg.walk():
                if "attachment" in str(part.get("Content-Disposition", "")).lower():
                    filename = part.get_filename()
                    if filename and filename.lower().endswith(".xlsx"):
                        filename = email.utils.collapse_rfc2231_value(filename)
                        safe_filename = filename.replace('/', '').replace('\\', '')
                        filepath = os.path.abspath(os.path.join(ATTACH_DIR, safe_filename))
                        with open(filepath, "wb") as f:
                            f.write(part.get_payload(decode=True))
                        attachments.append(filepath)

        for filepath in attachments:
            try:
                wb = load_workbook(filepath)
                for sheet in wb.worksheets:
                    for row in sheet.iter_rows(min_row=2, values_only=False):
                        for cell in row:
                            cell_value = clean_cell_value(cell.value)
                            for category, keywords in category_keywords.items():
                                if any(kw in cell_value for kw in keywords):
                                    c.execute("INSERT INTO matches (category, subject, sender, date, file_path, sheet, cell, cell_value) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                                              (category, subject, sender, date, filepath, sheet.title, cell.coordinate, str(cell.value)))
                                    conn.commit()
                                    break
            except Exception as e:
                print(f"Error reading {filepath}: {e}")

    conn.close()
    mail.logout()
    update_status("\u2705 Done! All data saved to database.")
    run_button.config(state="normal")
    run_button.config(text="\u25B6 Run Email Processor")
    progress_bar['value'] = 100
