from tkinter import *
from tkinter import ttk, messagebox
import threading
import sqlite3
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from db import init_db
from mail_parser import process_emails

DB_PATH = "matches.db"
category_keywords = {}

def add_category():
    cat = category_entry.get().strip()
    keys = keywords_entry.get().strip().lower().split(",")
    keys = [k.strip() for k in keys if k.strip()]
    if not cat or not keys:
        messagebox.showwarning("Invalid Input", "Please enter both category and keywords.")
        return
    category_keywords[cat] = keys
    listbox.insert(END, f"{cat}: {', '.join(keys)}")
    category_entry.delete(0, END)
    keywords_entry.delete(0, END)

def update_status(text):
    status_label.config(text=text)

def run_processor():
    if not category_keywords:
        messagebox.showerror("Error", "Add at least one category and keywords first.")
        return
    run_button.config(state="disabled")
    run_button.config(text="Processing...")
    progress_bar['value'] = 0
    update_status("Starting email scan...")
    threading.Thread(target=lambda: process_emails(category_keywords, update_status, progress_bar, root, run_button)).start()

def open_viewer_window():
    viewer = Toplevel(root)
    viewer.title("Browse Matches by Category")
    viewer.geometry("900x500")

    Label(viewer, text="Select Category:").pack()
    selected_category = StringVar(viewer)
    selected_category.set("Choose...")

    dropdown = OptionMenu(viewer, selected_category, *get_categories_from_db(), command=lambda cat: show_category_data(cat, viewer))
    dropdown.pack(pady=5)

def get_categories_from_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT DISTINCT category FROM matches")
    categories = [row[0] for row in c.fetchall()]
    conn.close()
    return categories

def show_category_data(category, parent):
    for widget in parent.winfo_children():
        if isinstance(widget, ttk.Treeview):
            widget.destroy()

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT subject, sender, date, file_path, sheet, cell, cell_value FROM matches WHERE category = ?", (category,))
    rows = c.fetchall()
    conn.close()

    tree = ttk.Treeview(parent, columns=("Subject", "Sender", "Date", "File", "Sheet", "Cell", "Value"), show='headings', height=20)
    for col in tree["columns"]:
        tree.heading(col, text=col)
        tree.column(col, width=120, anchor="w")
    for row in rows:
        tree.insert("", END, values=row)
    tree.pack(pady=10)

    Button(parent, text=f"\U0001F4E4 Export '{category}' to Excel", bg="skyblue", command=lambda: export_single_category(category)).pack(pady=5)

def export_single_category(category):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT subject, sender, date, file_path, sheet, cell, cell_value FROM matches WHERE category = ?", (category,))
    rows = c.fetchall()
    conn.close()
    if not rows:
        messagebox.showinfo("No Data", f"No matches found for category: {category}")
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Matches"
    ws.append(["Subject", "Sender", "Date", "Excel File Path", "Sheet", "Cell", "Cell Value"])
    for row in rows:
        ws.append(row)
    for col in ws.columns:
        max_length = max((len(str(cell.value)) for cell in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 5
    wb.save(f"{category}.xlsx")
    messagebox.showinfo("Exported", f"\u2705 Exported to {category}.xlsx")

def export_all():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT DISTINCT category FROM matches")
    categories = [row[0] for row in c.fetchall()]

    for category in categories:
        export_single_category(category)

# === GUI SETUP ===
init_db()
root = Tk()
root.title("Email Excel Parser with SQL")
root.geometry("540x520")

Label(root, text="Category Name").pack()
category_entry = Entry(root, width=50)
category_entry.pack()

Label(root, text="Keywords (comma-separated)").pack()
keywords_entry = Entry(root, width=50)
keywords_entry.pack()

Button(root, text="➕ Add Category", command=add_category).pack(pady=5)

scrollbar = Scrollbar(root)
scrollbar.pack(side="right", fill="y")
listbox = Listbox(root, width=60, yscrollcommand=scrollbar.set)
listbox.pack(pady=10)
scrollbar.config(command=listbox.yview)

run_button = Button(root, text="▶ Run Email Processor", bg="#4CAF50", fg="white", command=run_processor)
run_button.pack(pady=5)

Button(root, text="👁 View Matches", bg="gray", fg="white", command=open_viewer_window).pack(pady=5)
Button(root, text="📄 Export All to Excel", bg="teal", fg="white", command=export_all).pack(pady=5)

progress_bar = ttk.Progressbar(root, orient="horizontal", length=400, mode="determinate")
progress_bar.pack(pady=5)

status_label = Label(root, text="", fg="blue")
status_label.pack()

root.mainloop()
